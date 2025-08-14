from flask import Blueprint, request, jsonify, send_file
from backend.db.connection import get_db_connection
from backend.db.company_tables import get_company_table_name
from backend.auth.jwt_utils import token_required
from mysql.connector import Error
import datetime
from flask import current_app
import io
import csv
import tempfile
from openpyxl import load_workbook, Workbook
import re

calls_bp = Blueprint('calls', __name__)

# Helpers to normalize and validate Indian mobile numbers
def normalize_indian_mobile(raw: str) -> str:
    """Return a 10-digit Indian mobile number by removing optional prefixes and non-digits.
    Accepts formats like +91XXXXXXXXXX, 0XXXXXXXXXX, 91XXXXXXXXXX, and plain 10 digits.
    """
    if raw is None:
        return ''
    digits = re.sub(r'\D', '', str(raw))
    # Strip leading 0 (0 + 10 digits)
    if len(digits) == 11 and digits.startswith('0'):
        digits = digits[1:]
    # Strip country code 91 (91 + 10 digits)
    if len(digits) == 12 and digits.startswith('91'):
        digits = digits[2:]
    return digits

def is_valid_indian_mobile(ten_digits: str) -> bool:
    """Basic validation: exactly 10 digits and starts with 6/7/8/9."""
    return len(ten_digits) == 10 and ten_digits[0] in ('6', '7', '8', '9') and ten_digits.isdigit()

@calls_bp.route('/api/calls', methods=['GET'])
@token_required
def get_calls(current_user_id):
    # ... (move the full get_calls logic here from app.py)
    try:
        print("=== DEBUG: Calls endpoint called ===")
        # Get role and agent_number from JWT
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]
            except IndexError:
                return jsonify({'message': 'Token format invalid'}), 401
        import jwt
        data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=['HS256'])
        role = data.get('role', 'admin')
        agent_number = data.get('agent_number', None)
        company_id = data.get('company_id', None)

        # Date filter
        from_date = request.args.get('from')
        to_date = request.args.get('to')

        connection = get_db_connection()
        # connection = get_cdr_db_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        cursor = connection.cursor(dictionary=True)

        # Ensure recordings column is LONGBLOB
        try:
            check_cursor = connection.cursor(dictionary=True)
            check_cursor.execute(
                """
                SELECT DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'calls' AND COLUMN_NAME = 'recordings'
                """
            )
            col = check_cursor.fetchone()
            if col and col.get('DATA_TYPE', '').lower() != 'longblob':
                fix_cursor = connection.cursor()
                fix_cursor.execute("ALTER TABLE calls MODIFY recordings LONGBLOB")
                connection.commit()
        except Exception:
            pass

        # Enforce company isolation and status restrictions
        # If admin/agent, ensure company is Active for inbound; block outbound/export based on status on frontend via flags
        company_status = 'Active'
        payment_status = 'Paid'
        if company_id and role in ('admin', 'agent'):
            st = connection.cursor(dictionary=True)
            st.execute("SELECT status, payment_status FROM companies WHERE id = %s", (company_id,))
            row = st.fetchone()
            if row:
                company_status = row.get('status', 'Active')
                payment_status = row.get('payment_status', 'Paid')
            # If Fully Close, deny access entirely
            if company_status == 'Fully Close' or payment_status == 'Unpaid':
                return jsonify({'message': 'Access disabled for this company'}), 403

        # Build base query, using per-company calls table when company_id is present
        where_clauses = []
        params = []
        table_alias = 'c'
        if role in ('admin', 'agent') and company_id:
            calls_table = get_company_table_name('calls', int(company_id))
            base_query = (
                f"SELECT {table_alias}.id, {table_alias}.agent_number, {table_alias}.customer_number, {table_alias}.duration, {table_alias}.call_status, {table_alias}.timestamp, "
                f"{table_alias}.remarks, {table_alias}.name, {table_alias}.remarks_status, ({table_alias}.recordings IS NOT NULL) AS has_recording, {table_alias}.alternative_numbers, "
                f"{table_alias}.meeting_datetime, {table_alias}.meeting_description FROM `{calls_table}` {table_alias}"
            )
        else:
            base_query = (
                f"SELECT {table_alias}.id, {table_alias}.agent_number, {table_alias}.customer_number, {table_alias}.duration, {table_alias}.call_status, {table_alias}.timestamp, "
                f"{table_alias}.remarks, {table_alias}.name, {table_alias}.remarks_status, ({table_alias}.recordings IS NOT NULL) AS has_recording, {table_alias}.alternative_numbers, "
                f"{table_alias}.meeting_datetime, {table_alias}.meeting_description FROM calls {table_alias}"
            )
        if role == 'agent' and agent_number:
            where_clauses.append(f'{table_alias}.agent_number = %s')
            params.append(agent_number)
        if from_date:
            where_clauses.append(f'{table_alias}.timestamp >= %s')
            params.append(from_date + ' 00:00:00')
        if to_date:
            where_clauses.append(f'{table_alias}.timestamp <= %s')
            params.append(to_date + ' 23:59:59')
        if where_clauses:
            base_query += ' WHERE ' + ' AND '.join(where_clauses)
        base_query += f' ORDER BY {table_alias}.timestamp DESC'
        cursor.execute(base_query, tuple(params))
        calls = cursor.fetchall()
        # Normalize timestamps and unify customer names within the same customer_number group
        latest_name_by_customer = {}
        for call in calls:
            if call['timestamp']:
                call['timestamp'] = call['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
            if call.get('meeting_datetime'):
                try:
                    call['meeting_datetime'] = call['meeting_datetime'].strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    pass
            cust = call.get('customer_number')
            name = (call.get('name') or '').strip()
            if cust:
                # Because we ordered DESC by timestamp, the first seen name per customer is the latest
                if cust not in latest_name_by_customer and name:
                    latest_name_by_customer[cust] = name
        for call in calls:
            cust = call.get('customer_number')
            if cust and latest_name_by_customer.get(cust):
                call['name'] = latest_name_by_customer[cust]
        return jsonify({'calls': calls}), 200
    except Exception as e:
        return jsonify({'message': f'Server error: {str(e)}'}), 500
    finally:
        try:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'connection' in locals() and connection:
                connection.close()
        except Exception as e:
            print(f"Error closing resources: {e}")

@calls_bp.route('/api/calls/<int:call_id>/custom', methods=['PUT'])
@token_required
def update_call_custom_fields(current_user_id, call_id):
    # ... (move the full update_call_custom_fields logic here from app.py)
    try:
        data = request.get_json()
        remarks = data.get('remarks', '')
        name = data.get('name', '')
        remarks_status = data.get('remarks_status', '')
        recordings = None  # recordings blob handled via separate endpoint
        alternative_numbers = data.get('alternative_numbers', '')
        meeting_datetime = data.get('meeting_datetime')
        meeting_description = data.get('meeting_description')

        connection = get_db_connection()
        # connection = get_cdr_db_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        cursor = connection.cursor(dictionary=True)

        # Fetch current values
        # Determine per-company table from token
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]
            except IndexError:
                return jsonify({'message': 'Token format invalid'}), 401
        import jwt
        jwt_data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=['HS256'])
        company_id = jwt_data.get('company_id')
        if company_id:
            calls_table = get_company_table_name('calls', int(company_id))
            cursor.execute(f"SELECT customer_number, remarks, name, remarks_status, alternative_numbers FROM `{calls_table}` WHERE id = %s", (call_id,))
        else:
            cursor.execute("SELECT customer_number, remarks, name, remarks_status, alternative_numbers FROM calls WHERE id = %s", (call_id,))
        call = cursor.fetchone()
        if not call:
            return jsonify({'message': 'Call not found'}), 404

        now = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        def concat(old, new):
            old = old or ''
            if new:
                return (old + ' | ' if old else '') + f"{new} <span style='font-size:10px;color:gray;'>[{now}]</span>"
            return old

        updated_remarks = concat(call.get('remarks', ''), remarks)
        # For name: single canonical value per customer. If provided, replace and propagate to all calls for this customer.
        updated_name = call.get('name', '')
        if name and name.strip():
            updated_name = name.strip()
        updated_remarks_status = concat(call.get('remarks_status', ''), remarks_status)
        updated_alternative_numbers = alternative_numbers if alternative_numbers else call.get('alternative_numbers', '')
        
        if company_id:
            cursor.execute(
                f"""
                UPDATE `{calls_table}` SET remarks=%s, name=%s, remarks_status=%s, alternative_numbers=%s,
                       meeting_datetime=%s, meeting_description=%s
                WHERE id=%s
                """,
                (
                    updated_remarks,
                    updated_name,
                    updated_remarks_status,
                    updated_alternative_numbers,
                    meeting_datetime,
                    meeting_description,
                    call_id,
                ),
            )
        else:
            cursor.execute(
                """
                UPDATE calls SET remarks=%s, name=%s, remarks_status=%s, alternative_numbers=%s,
                       meeting_datetime=%s, meeting_description=%s
                WHERE id=%s
                """,
                (
                    updated_remarks,
                    updated_name,
                    updated_remarks_status,
                    updated_alternative_numbers,
                    meeting_datetime,
                    meeting_description,
                    call_id,
                ),
            )
        # If name changed, propagate to all rows for that customer
        if name and name.strip():
            if company_id:
                cursor.execute(
                    f"UPDATE `{calls_table}` SET name=%s WHERE customer_number=%s",
                    (updated_name, call['customer_number']),
                )
            else:
                cursor.execute(
                    "UPDATE calls SET name=%s WHERE customer_number=%s",
                    (updated_name, call['customer_number']),
                )
        connection.commit()
        return jsonify({'message': 'Custom fields updated successfully'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()  # Print full stack trace to console
        print("Updated Remarks:", updated_remarks)
        print("Updated Name:", updated_name)
        print("Updated Remarks Status:", updated_remarks_status)
        # recordings handled separately
        print("Updated Alternative Numbers:", updated_alternative_numbers)
        print("Call ID:", call_id)
        return jsonify({'message': f'Server error: {str(e)}'}), 500
        
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@calls_bp.route('/api/calls/<int:call_id>/recording', methods=['GET'])
@token_required
def download_call_recording(current_user_id, call_id):
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        cursor = connection.cursor()
        # Determine company
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]
            except IndexError:
                return jsonify({'message': 'Token format invalid'}), 401
        import jwt
        jwt_data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=['HS256'])
        company_id = jwt_data.get('company_id')
        if company_id:
            calls_table = get_company_table_name('calls', int(company_id))
            cursor.execute(f"SELECT recordings FROM `{calls_table}` WHERE id=%s", (call_id,))
        else:
            cursor.execute("SELECT recordings FROM calls WHERE id=%s", (call_id,))
        row = cursor.fetchone()
        if not row or row[0] is None:
            return jsonify({'message': 'Recording not found'}), 404
        blob = row[0]
        # Try to infer audio MIME type from file header
        mime = 'audio/mpeg'
        try:
            header = bytes(blob[:4]) if isinstance(blob, (bytes, bytearray)) else b''
            if header.startswith(b'RIFF'):
                mime = 'audio/wav'
            elif header.startswith(b'OggS'):
                mime = 'audio/ogg'
            elif header[:3] == b'ID3' or header[:2] == b'\xff\xfb':
                mime = 'audio/mpeg'
        except Exception:
            pass
        resp = send_file(io.BytesIO(blob), mimetype=mime)
        try:
            resp.headers['Accept-Ranges'] = 'bytes'
        except Exception:
            pass
        return resp
    except Exception as e:
        return jsonify({'message': f'Server error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()


@calls_bp.route('/api/calls/upload-corrections', methods=['POST'])
@token_required
def upload_call_corrections(current_user_id):
    """Accept JSON list of corrected rows and insert them.
    Each item must include: agent_number, customer_number, name, remarks
    """
    try:
        body = request.get_json() or {}
        rows = body.get('rows', [])
        # Decode token to get company_id
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(' ')[1]
            except IndexError:
                return jsonify({'message': 'Token format invalid'}), 401
        import jwt
        jwt_data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=['HS256'])
        company_id = jwt_data.get('company_id')
        if not isinstance(rows, list) or not rows:
            return jsonify({'message': 'No rows provided'}), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        cursor = connection.cursor()

        inserted = 0
        errors = []
        row_number = 1
        for item in rows:
            row_number += 1
            effective_agent_number = str(item.get('agent_number', '')).strip()
            customer_number_raw = str(item.get('customer_number', '')).strip()
            name = str(item.get('name', '')).strip()
            remarks = str(item.get('remarks', '')).strip()

            if not effective_agent_number or not customer_number_raw:
                errors.append({
                    'row': row_number,
                    'agent_number': effective_agent_number,
                    'customer_number': customer_number_raw,
                    'name': name,
                    'remarks': remarks,
                    'reason': 'Missing required agent_number or customer_number'
                })
                continue

            digits_only = normalize_indian_mobile(customer_number_raw)
            if not is_valid_indian_mobile(digits_only):
                errors.append({
                    'row': row_number,
                    'agent_number': effective_agent_number,
                    'customer_number': customer_number_raw,
                    'name': name,
                    'remarks': remarks,
                    'reason': 'Invalid Indian mobile number'
                })
                continue

            calls_table = get_company_table_name('calls', int(company_id))
            cursor.execute(
                f"""
                INSERT INTO `{calls_table}` (agent_number, customer_number, duration, call_status, timestamp, remarks, name, remarks_status)
                VALUES (%s, %s, %s, %s, NOW(), %s, %s, %s)
                """,
                (effective_agent_number, digits_only, 0, 'Uploaded', remarks, name, '')
            )
            inserted += 1

        connection.commit()
        return jsonify({
            'message': f'Processed corrections: {inserted} success, {len(errors)} errors',
            'success_count': inserted,
            'error_count': len(errors),
            'errors': errors
        }), 200
    except Exception as e:
        return jsonify({'message': f'Server error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@calls_bp.route('/api/calls/<int:call_id>/recording', methods=['PUT'])
@token_required
def upload_call_recording(current_user_id, call_id):
    try:
        if 'file' not in request.files:
            return jsonify({'message': 'No file provided'}), 400
        f = request.files['file']
        data = f.read()
        connection = get_db_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        cursor = connection.cursor()
        # Determine company
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(' ')[1]
            except IndexError:
                return jsonify({'message': 'Token format invalid'}), 401
        import jwt
        jwt_data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=['HS256'])
        company_id = jwt_data.get('company_id')
        if company_id:
            calls_table = get_company_table_name('calls', int(company_id))
            # Ensure column type
            try:
                cursor.execute(f"ALTER TABLE `{calls_table}` MODIFY recordings LONGBLOB")
                connection.commit()
            except Exception:
                pass
            cursor.execute(f"UPDATE `{calls_table}` SET recordings=%s WHERE id=%s", (data, call_id))
        else:
            try:
                cursor.execute("ALTER TABLE calls MODIFY recordings LONGBLOB")
                connection.commit()
            except Exception:
                pass
            cursor.execute("UPDATE calls SET recordings=%s WHERE id=%s", (data, call_id))
        connection.commit()
        return jsonify({'message': 'Recording uploaded'}), 200
    except Exception as e:
        return jsonify({'message': f'Server error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@calls_bp.route('/api/calls/upload-template', methods=['GET'])
@token_required
def download_calls_template(current_user_id):
    """Provide an Excel template for uploading call details.

    Columns included by default: agent_number, customer, name, remarks
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'CallsTemplate'
    # Header per requirement
    ws.append(['agent_number', 'customer', 'name', 'remarks'])
    # Example row (optional)
    ws.append(['A001', '9876543210', 'Customer Name', 'Optional initial remark'])
    tmp = io.BytesIO()
    wb.save(tmp)
    tmp.seek(0)
    return send_file(
        tmp,
        as_attachment=True,
        download_name='calls_upload_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

@calls_bp.route('/api/calls/upload', methods=['POST'])
@token_required
def upload_calls_excel(current_user_id):
    """Admin uploads calls; rows are created with status 'Uploaded'.
    Expected form fields: file (Excel .xlsx). Optional 'agent_number' for fallback.
    Accepts spreadsheet columns: agent_number (per row), customer/customer_number, name, remarks.
    If per-row agent_number is present, it overrides the form agent_number for that row.
    """
    try:
        # Ensure admin
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(' ')[1]
            except IndexError:
                return jsonify({'message': 'Token format invalid'}), 401
        import jwt
        data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=['HS256'])
        role = data.get('role', 'admin')
        company_id = data.get('company_id')
        if role != 'admin':
            return jsonify({'message': 'Unauthorized'}), 403

        # agent_number in form is optional now; used as fallback if row is missing it
        agent_number = request.form.get('agent_number', '').strip()
        if 'file' not in request.files:
            return jsonify({'message': 'Excel file is required'}), 400
        f = request.files['file']
        if not f.filename.lower().endswith('.xlsx'):
            return jsonify({'message': 'Please upload an .xlsx Excel file'}), 400

        wb = load_workbook(filename=io.BytesIO(f.read()))
        ws = wb.active

        connection = get_db_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        cursor = connection.cursor()

        inserted = 0
        errors = []

        # Build header index mapping (case-insensitive)
        header_row = next(ws.iter_rows(values_only=True))
        header_to_index = {}
        if header_row:
            for idx, header in enumerate(header_row):
                key = str(header or '').strip().lower()
                if key:
                    header_to_index[key] = idx

        # Resolve column indices with synonyms
        idx_agent = header_to_index.get('agent_number')
        idx_customer = (
            header_to_index.get('customer')
            if 'customer' in header_to_index
            else header_to_index.get('customer_number')
        )
        idx_name = header_to_index.get('name')
        idx_remarks = header_to_index.get('remarks')

        # Iterate data rows
        # Iterate from second row to keep the header reference intact
        first = True
        row_number = 1
        for row in ws.iter_rows(values_only=True):
            row_number += 1
            if first:
                first = False
                continue
            if not row:
                continue

            # Helper to safely get and normalize cell text
            def cell_text(index):
                if index is None or index >= len(row):
                    return ''
                value = row[index]
                return (value or '').strip() if isinstance(value, str) else str(value or '').strip()

            row_agent_number = cell_text(idx_agent) if idx_agent is not None else ''
            effective_agent_number = row_agent_number or agent_number
            customer_number = cell_text(idx_customer) if idx_customer is not None else cell_text(0)
            name = cell_text(idx_name) if idx_name is not None else cell_text(1)
            remarks = cell_text(idx_remarks) if idx_remarks is not None else cell_text(2)

            # Validate required fields
            if not effective_agent_number or not customer_number:
                errors.append({
                    'row': row_number,
                    'agent_number': effective_agent_number,
                    'customer_number': customer_number,
                    'name': name,
                    'remarks': remarks,
                    'reason': 'Missing required agent_number or customer_number'
                })
                continue

            # Normalize and validate Indian mobile number (10 digits, allow +91/0 prefixes)
            digits_only = normalize_indian_mobile(customer_number)
            if not is_valid_indian_mobile(digits_only):
                errors.append({
                    'row': row_number,
                    'agent_number': effective_agent_number,
                    'customer_number': customer_number,
                    'name': name,
                    'remarks': remarks,
                    'reason': 'Invalid Indian mobile number'
                })
                continue

            # Ensure agent exists in this organization
            agents_table = get_company_table_name('agents', int(company_id))
            cursor.execute(f"SELECT 1 FROM `{agents_table}` WHERE agent_number = %s", (effective_agent_number,))
            if cursor.fetchone() is None:
                errors.append({
                    'row': row_number,
                    'agent_number': effective_agent_number,
                    'customer_number': customer_number,
                    'name': name,
                    'remarks': remarks,
                    'reason': 'Agent number not found in this organization'
                })
                continue

            calls_table = get_company_table_name('calls', int(company_id))
            cursor.execute(
                f"""
                INSERT INTO `{calls_table}` (agent_number, customer_number, duration, call_status, timestamp, remarks, name, remarks_status)
                VALUES (%s, %s, %s, %s, NOW(), %s, %s, %s)
                """,
                (effective_agent_number, digits_only, 0, 'Uploaded', remarks, name, '')
            )
            inserted += 1
        connection.commit()
        return jsonify({
            'message': f'Processed file: {inserted} success, {len(errors)} errors',
            'success_count': inserted,
            'error_count': len(errors),
            'errors': errors
        }), 201
    except Exception as e:
        return jsonify({'message': f'Server error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@calls_bp.route('/api/calls/<int:call_id>/alternative-numbers', methods=['PUT'])
@token_required
def update_alternative_numbers(current_user_id, call_id):
    # ... (move the full update_alternative_numbers logic here from app.py)
    try:
        data = request.get_json()
        alternative_numbers = data.get('alternative_numbers', '')

        connection = get_db_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        cursor = connection.cursor(dictionary=True)

        # Update alternative numbers
        # Determine company
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]
            except IndexError:
                return jsonify({'message': 'Token format invalid'}), 401
        import jwt
        jwt_data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=['HS256'])
        company_id = jwt_data.get('company_id')
        if company_id:
            calls_table = get_company_table_name('calls', int(company_id))
            cursor.execute(
                f"""
                UPDATE `{calls_table}` SET alternative_numbers=%s WHERE id=%s
                """,
                (alternative_numbers, call_id)
            )
        else:
            cursor.execute(
                """
                UPDATE calls SET alternative_numbers=%s WHERE id=%s
                """,
                (alternative_numbers, call_id)
            )
        connection.commit()
        return jsonify({'message': 'Alternative numbers updated successfully'}), 200
    except Exception as e:
        return jsonify({'message': f'Server error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()